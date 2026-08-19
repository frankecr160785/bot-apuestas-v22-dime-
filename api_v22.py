import os
import sys
import time
import argparse
import requests
from datetime import datetime
from openpyxl import load_workbook

BASE_URL = "https://v3.football.api-sports.io"

ALLOWED_COUNTRIES = [
    "England", "Spain", "Italy", "Germany", "France",
    "Netherlands", "Portugal", "Switzerland", "Sweden",
    "Turkey", "Finland", "Iceland",
    "Brazil", "Argentina", "Colombia", "Mexico", "USA",
    "China", "Japan", "South-Korea"
]

# Países prioritarios para el motor.
GOAL_FOCUS_COUNTRIES = {"Finland", "Iceland"}

PREFERRED_LEAGUES = {
    "England": ["Premier League", "Championship"],
    "Spain": ["La Liga", "Segunda División"],
    "Italy": ["Serie A", "Serie B"],
    "Germany": ["Bundesliga", "2. Bundesliga"],
    "France": ["Ligue 1", "Ligue 2"],
    "Netherlands": ["Eredivisie", "Eerste Divisie"],
    "Portugal": ["Primeira Liga"],
    "Switzerland": ["Super League", "Challenge League"],
    "Sweden": ["Allsvenskan", "Superettan"],
    "Turkey": ["Süper Lig", "1. Lig"],
    "Finland": ["Veikkausliiga", "Ykkösliiga", "Ykkönen"],
    "Iceland": ["Úrvalsdeild", "1. Deild", "2. Deild"],
    "Brazil": ["Serie A", "Serie B"],
    "Argentina": ["Liga Profesional Argentina", "Primera Nacional"],
    "Colombia": ["Primera A", "Primera B"],
    "Mexico": ["Liga MX", "Liga de Expansión MX"],
    "USA": ["Major League Soccer", "USL Championship"],
    "China": ["Super League", "League One"],
    "Japan": ["J1 League", "J2 League", "J3 League"],
    "South-Korea": ["K League 1", "K League 2", "K3 League"]
}

# Variantes para evitar que un cambio de nombre en API-Football deje fuera Islandia.
LEAGUE_ALIASES = {
    "Iceland": {
        "Úrvalsdeild": {"Úrvalsdeild", "Besta deild karla", "Besta-deild karla"},
        "1. Deild": {"1. Deild", "1. deild karla", "1. deild"},
        "2. Deild": {"2. Deild", "2. deild karla", "2. deild"},
    },
    "Finland": {
        "Veikkausliiga": {"Veikkausliiga"},
        "Ykkösliiga": {"Ykkösliiga", "Ykkosliiga"},
        "Ykkönen": {"Ykkönen", "Ykkonen", "Ykkönen -"},
    }
}


class RateLimiter:
    def __init__(self, interval=6.5):
        self.interval = interval
        self.last = 0.0

    def wait(self):
        delay = self.interval - (time.time() - self.last)
        if delay > 0:
            time.sleep(delay)
        self.last = time.time()


def api_get(session, limiter, path, params, key, retries=3):
    for attempt in range(retries):
        limiter.wait()
        r = session.get(
            BASE_URL + path,
            headers={"x-apisports-key": key},
            params=params,
            timeout=45
        )

        if r.status_code == 429:
            wait = 15 * (attempt + 1)
            print(f"429: esperando {wait}s...")
            time.sleep(wait)
            continue

        r.raise_for_status()
        data = r.json()

        if data.get("errors"):
            raise RuntimeError(str(data["errors"]))

        return data.get("response", [])

    raise RuntimeError("Límite de API alcanzado después de varios intentos")


def get_leagues(session, limiter, key, country, season):
    return api_get(
        session, limiter, "/leagues",
        {"country": country, "season": season},
        key
    )


def get_fixtures(session, limiter, key, league_id, season, date_s):
    return api_get(
        session, limiter, "/fixtures",
        {
            "league": league_id,
            "season": season,
            "date": date_s,
            "timezone": "America/Lima"
        },
        key
    )


def get_stats(session, limiter, key, fixture_id):
    return api_get(
        session, limiter, "/fixtures/statistics",
        {"fixture": fixture_id},
        key
    )


def get_odds(session, limiter, key, fixture_id):
    return api_get(
        session, limiter, "/odds",
        {"fixture": fixture_id},
        key
    )


def get_odds_page(session, limiter, key, date_s, page=1):
    return api_get(
        session, limiter, "/odds",
        {"date": date_s, "page": page},
        key
    )


def get_fixture_details_bulk(session, limiter, key, fixture_ids):
    """Recupera detalles de hasta 20 fixtures por llamada."""
    details = []
    ids = [str(x) for x in fixture_ids if x]
    for start in range(0, len(ids), 20):
        chunk = ids[start:start + 20]
        if not chunk:
            continue
        response = api_get(
            session,
            limiter,
            "/fixtures",
            {"ids": "-".join(chunk)},
            key
        )
        details.extend(response)
    return details


def normalize_name(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def league_is_preferred(country, league_name):
    name = normalize_name(league_name)

    for preferred in PREFERRED_LEAGUES.get(country, []):
        if name == normalize_name(preferred):
            return True

    # Variantes específicas, especialmente Islandia/Finlandia.
    for canonical, aliases in LEAGUE_ALIASES.get(country, {}).items():
        if name in {normalize_name(x) for x in aliases}:
            return True

    return False


def extract_stat(stats_response, stat_name):
    """ Busca una estadística dentro de la respuesta de /fixtures/statistics y devuelve el valor de HOME/AWAY. """
    result = {"home": None, "away": None}

    for team_block in stats_response:
        team = team_block.get("team", {})
        team_name = team.get("name", "")
        side = "home" if team_block.get("_side") == "home" else "away"

        for item in team_block.get("statistics", []):
            if normalize_name(item.get("type")) == normalize_name(stat_name):
                result[side] = item.get("value")

    return result


def prepare_stats_rows(fixture, stats_response):
    teams = fixture.get("teams", {})
    home = teams.get("home", {}).get("name", "")
    away = teams.get("away", {}).get("name", "")
    fixture_id = fixture.get("fixture", {}).get("id")
    date_s = fixture.get("fixture", {}).get("date", "")[:10]

    # API-Football no entrega el lado en cada bloque; lo asignamos
    # comparando el team.id con home/away.
    home_id = teams.get("home", {}).get("id")
    away_id = teams.get("away", {}).get("id")

    rows = []
    for block in stats_response:
        team_id = block.get("team", {}).get("id")
        side = "home" if team_id == home_id else "away" if team_id == away_id else ""

        values = {}
        for item in block.get("statistics", []):
            values[normalize_name(item.get("type"))] = item.get("value")

        rows.append({
            "fixture_id": fixture_id,
            "date": date_s,
            "league": fixture.get("league", {}).get("name", ""),
            "country": fixture.get("_country", ""),
            "home": home,
            "away": away,
            "side": side,
            "team": block.get("team", {}).get("name", ""),
            "shots_on_goal": values.get("shots on goal"),
            "shots_off_goal": values.get("shots off goal"),
            "total_shots": values.get("total shots"),
            "blocked_shots": values.get("blocked shots"),
            "corners": values.get("corner kicks"),
            "fouls": values.get("fouls"),
            "offsides": values.get("offsides"),
            "ball_possession": values.get("ball possession"),
            "yellow_cards": values.get("yellow cards"),
            "red_cards": values.get("red cards"),
            "passes": values.get("total passes"),
            "passes_accurate": values.get("passes accurate"),
            "passes_pct": values.get("passes %"),
            "expected_goals": values.get("expected goals"),
            "goals_prevented": values.get("goals prevented"),
        })

    return rows


def flatten_odds(fixture, odds_response):
    """ Convierte la respuesta de /odds en filas simples. Conserva bookmaker, mercado, selección y cuota. """
    rows = []
    teams = fixture.get("teams", {})
    home = teams.get("home", {}).get("name", "")
    away = teams.get("away", {}).get("name", "")
    fixture_id = fixture.get("fixture", {}).get("id")
    date_s = fixture.get("fixture", {}).get("date", "")[:10]

    for bookmaker_block in odds_response:
        bookmaker = bookmaker_block.get("bookmaker", {})
        bookmaker_name = bookmaker.get("name", "")

        for bet in bookmaker_block.get("bets", []):
            market = bet.get("name", "")
            for value in bet.get("values", []):
                rows.append({
                    "fixture_id": fixture_id,
                    "date": date_s,
                    "league": fixture.get("league", {}).get("name", ""),
                    "country": fixture.get("_country", ""),
                    "home": home,
                    "away": away,
                    "bookmaker": bookmaker_name,
                    "market": market,
                    "selection": value.get("value", ""),
                    "odd": value.get("odd"),
                })

    return rows


def ensure_sheet(wb, title, headers):
    if title in wb.sheetnames:
        ws = wb[title]
        # No borrar información previa: sólo agregamos encabezados si está vacío.
        if ws.max_row == 1 and all(ws.cell(1, c).value is None for c in range(1, len(headers) + 1)):
            for c, header in enumerate(headers, 1):
                ws.cell(1, c).value = header
        elif ws.max_row == 1:
            existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if not any(existing):
                for c, header in enumerate(headers, 1):
                    ws.cell(1, c).value = header
        return ws

    ws = wb.create_sheet(title)
    for c, header in enumerate(headers, 1):
        ws.cell(1, c).value = header
    return ws


def existing_keys(ws, key_cols):
    keys = set()
    if ws.max_row < 2:
        return keys

    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    if not all(k in header for k in key_cols):
        return keys

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = tuple(row[header[k] - 1] for k in key_cols)
        if any(v not in (None, "") for v in key):
            keys.add(tuple(str(v) if v is not None else "" for v in key))
    return keys


def append_dict_rows(ws, rows):
    if not rows:
        return 0

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not any(headers):
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h

    header_map = {h: i + 1 for i, h in enumerate(headers) if h}
    added = 0

    for item in rows:
        ws.append([
            item.get(h, "") if h else ""
            for h in headers
        ])
        added += 1

    return added


def fill_excel(xlsx, fixtures, stats_rows, odds_rows):
    wb = load_workbook(xlsx)

    # Hoja principal existente.
    if "BASE_DATOS" not in wb.sheetnames:
        ws_base = wb.create_sheet("BASE_DATOS")
        base_headers = [
            "Fecha", "Liga", "Local", "Condición", "Visitante",
            "fixture_id", "País"
        ]
        for c, h in enumerate(base_headers, 1):
            ws_base.cell(1, c).value = h
    else:
        ws_base = wb["BASE_DATOS"]

    # El Excel V2.2 usa FECHA/LIGA/EQUIPO/CONDICIÓN/RIVAL.
    # Reparamos las filas vacías creadas por V2.3.1 y luego evitamos duplicados.
    headers = [ws_base.cell(1, c).value for c in range(1, ws_base.max_column + 1)]
    hmap = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}

    def set_cell(row_idx, names, value):
        for name in names:
            if name in hmap:
                ws_base.cell(row_idx, hmap[name] + 1).value = value
                return True
        return False

    def row_value(row_idx, names):
        for name in names:
            if name in hmap:
                return ws_base.cell(row_idx, hmap[name] + 1).value
        return None

    # Identificar filas completamente vacías de equipo/rival que dejó la ejecución anterior.
    blank_rows = []
    for r in range(2, ws_base.max_row + 1):
        team = row_value(r, ["equipo", "team", "local", "home"])
        rival = row_value(r, ["rival", "visitante", "away"])
        if team in (None, "") and rival in (None, ""):
            row_date = row_value(r, ["fecha", "date"])
            if str(row_date or "")[:10] == str(args_date)[:10]:
                blank_rows.append(r)

    # Orden estable para que la reparación sea reproducible.
    fixtures_for_base = sorted(
        fixtures,
        key=lambda f: (
            f.get("_country", ""),
            f.get("league", {}).get("name", ""),
            f.get("fixture", {}).get("id", 0)
        )
    )

    repaired = 0
    used_blank = set()
    for f in fixtures_for_base:
        if not blank_rows:
            break
        fx = f.get("fixture", {})
        teams = f.get("teams", {})
        league = f.get("league", {})
        date_s = str(fx.get("date", "") or "")[:10] or args_date
        home = teams.get("home", {}).get("name", "")
        away = teams.get("away", {}).get("name", "")
        league_name = league.get("name", "")
        fid = fx.get("id")
        r = blank_rows.pop(0)
        used_blank.add(r)

        set_cell(r, ["fecha", "date"], date_s)
        set_cell(r, ["liga", "league"], league_name)
        set_cell(r, ["equipo", "team", "local", "home"], home)
        set_cell(r, ["condición", "condicion", "side"], "LOCAL")
        set_cell(r, ["rival", "visitante", "away"], away)
        if "observaciones" in hmap:
            ws_base.cell(r, hmap["observaciones"] + 1).value = (
                f"API-Football | fixture={fid} | país={f.get('_country','')}"
            )
        repaired += 1

    # Clave real del esquema V2.2.
    existing = existing_keys(ws_base, ["FECHA", "EQUIPO", "RIVAL", "CONDICIÓN"])
    added_base = repaired

    for f in fixtures_for_base:
        fx = f.get("fixture", {})
        teams = f.get("teams", {})
        league = f.get("league", {})
        date_s = str(fx.get("date", "") or "")[:10] or args_date
        home = teams.get("home", {}).get("name", "")
        away = teams.get("away", {}).get("name", "")
        league_name = league.get("name", "")
        country = f.get("_country", "")
        fid = fx.get("id")

        rows_to_add = [
            (home, "LOCAL", away),
            (away, "VISITANTE", home),
        ]

        for team_name, condition, rival_name in rows_to_add:
            key = (str(date_s), str(team_name), str(rival_name), str(condition))
            if key in existing:
                continue

            row = [None] * len(headers)
            values = {
                "fecha": date_s,
                "date": date_s,
                "liga": league_name,
                "league": league_name,
                "equipo": team_name,
                "team": team_name,
                "local": home,
                "home": home,
                "condición": condition,
                "condicion": condition,
                "side": condition,
                "rival": rival_name,
                "visitante": away,
                "away": away,
                "observaciones": f"API-Football | fixture={fid} | país={country}",
            }
            for name, value in values.items():
                if name in hmap:
                    row[hmap[name]] = value

            ws_base.append(row)
            existing.add(key)
            added_base += 1

    print(f"BASE_DATOS reparados/nuevos: {added_base}")

    # Estadísticas completas.
    stats_headers = [
        "fixture_id", "date", "league", "country", "home", "away",
        "side", "team", "shots_on_goal", "shots_off_goal", "total_shots",
        "blocked_shots", "corners", "fouls", "offsides", "ball_possession",
        "yellow_cards", "red_cards", "passes", "passes_accurate",
        "passes_pct", "expected_goals", "goals_prevented"
    ]

    ws_stats = ensure_sheet(wb, "API_STATS", stats_headers)
    stats_existing = existing_keys(ws_stats, ["fixture_id", "team"])

    new_stats = []
    for row in stats_rows:
        k = (str(row.get("fixture_id", "")), str(row.get("team", "")))
        if k not in stats_existing:
            new_stats.append(row)
            stats_existing.add(k)

    added_stats = append_dict_rows(ws_stats, new_stats)

    # Cuotas.
    odds_headers = [
        "fixture_id", "date", "league", "country", "home", "away",
        "bookmaker", "market", "selection", "odd"
    ]

    ws_odds = ensure_sheet(wb, "API_ODDS", odds_headers)
    odds_existing = existing_keys(
        ws_odds,
        ["fixture_id", "bookmaker", "market", "selection"]
    )

    new_odds = []
    for row in odds_rows:
        k = (
            str(row.get("fixture_id", "")),
            str(row.get("bookmaker", "")),
            str(row.get("market", "")),
            str(row.get("selection", ""))
        )
        if k not in odds_existing:
            new_odds.append(row)
            odds_existing.add(k)

    added_odds = append_dict_rows(ws_odds, new_odds)

    # Hojas auxiliares para identificar fácilmente la procedencia.
    if "CONTROL_API" not in wb.sheetnames:
        ws_control = wb.create_sheet("CONTROL_API")
        control_headers = [
            "Ejecución", "Fecha consultada", "Partidos encontrados",
            "BASE_DATOS nuevos", "API_STATS nuevos", "API_ODDS nuevos",
            "Países prioritarios"
        ]
        for c, h in enumerate(control_headers, 1):
            ws_control.cell(1, c).value = h
    else:
        ws_control = wb["CONTROL_API"]

    ws_control.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        fixtures[0].get("fixture", {}).get("date", "")[:10] if fixtures else "",
        len(fixtures),
        added_base,
        added_stats,
        added_odds,
        ", ".join(sorted(GOAL_FOCUS_COUNTRIES))
    ])

    wb.save(xlsx)

    print(f"BASE_DATOS nuevos: {added_base}")
    print(f"API_STATS nuevos: {added_stats}")
    print(f"API_ODDS nuevos: {added_odds}")


def main():
    ap = argparse.ArgumentParser(description="Bot Apuestas V2.3.2 - optimizado para Excel V2.2")
    ap.add_argument(
        "--date",
        default=datetime.now().strftime("%Y-%m-%d")
    )
    ap.add_argument(
        "--season",
        type=int,
        default=datetime.now().year
    )
    ap.add_argument(
        "--xlsx",
        default="Excel_V2_2_Motor_Automatico.xlsx"
    )
    ap.add_argument(
        "--details",
        action="store_true"
    )
    ap.add_argument(
        "--max-fixtures",
        type=int,
        default=30
    )

    args = ap.parse_args()
    args_date = args.date

    key = os.getenv("API_FOOTBALL_KEY")
    if not key:
        print("ERROR: falta API_FOOTBALL_KEY")
        sys.exit(2)

    if not os.path.exists(args.xlsx):
        print(f"ERROR: no existe {args.xlsx}")
        sys.exit(3)

    session = requests.Session()
    limiter = RateLimiter()

    # V2.3.1:
    # Consultamos TODOS los partidos del día en una sola llamada y después
    # filtramos por país/liga. Esto evita que un cambio de temporada o nombre
    # de liga deje el motor con 0 partidos aunque la API sí tenga encuentros.
    print(f"Consultando partidos del día {args.date} / zona America-Lima...")

    fixtures = []

    try:
        all_fixtures = api_get(
            session,
            limiter,
            "/fixtures",
            {
                "date": args.date,
                "timezone": "America/Lima"
            },
            key
        )

        print(f"Partidos devueltos por API para la fecha: {len(all_fixtures)}")

        allowed = {normalize_name(c): c for c in ALLOWED_COUNTRIES}

        for f in all_fixtures:
            league = f.get("league", {})
            country_name = league.get("country", "")
            league_name = league.get("name", "")

            canonical_country = allowed.get(normalize_name(country_name))
            if not canonical_country:
                continue 
